import io
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# Константы
INCH_TO_CHAR = (9.85546875 / 0.82)
INCH_TO_CHAR_H = 72

# Цвета
GREEN_FILL = PatternFill(start_color="80FF80", end_color="80FF80", fill_type="solid")
RED_FILL = PatternFill(start_color="FF8080", end_color="FF8080", fill_type="solid")
GREY_FILL = PatternFill(start_color="DBDBDB", end_color="DBDBDB", fill_type="solid")
GREENY_FILL = PatternFill(start_color="c5e0b4", end_color="c5e0b4", fill_type="solid")
BLUE_FILL = PatternFill(start_color="9dc3e6", end_color="9dc3e6", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="ffff99", end_color="ffff99", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="ffe699", end_color="ffe699", fill_type="solid")
LIGHT_PURPLE_FILL = PatternFill(start_color="b4c7e7", end_color="b4c7e7", fill_type="solid")
BEIGE_FILL = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")


def style_cell(cell, size=11, bold=False, horiz='center', vert='center'):
    cell.font = Font(name='Times New Roman', size=size, bold=bold)
    cell.alignment = Alignment(horizontal=horiz, vertical=vert, wrap_text=True)


def get_dynamic_schema(data):
    info_keys, vdpg_keys = set(), set()
    cats = ["Депутаты Законодательных собраний регионов",
            "Депутаты административных центров регионов",
            "Депутаты муниципальных образований"]
    for cat in cats:
        for dep in data.get(cat, []):
            if "Посты по информационным ударам" in dep:
                info_keys.update(dep["Посты по информационным ударам"].keys())
            if "ВДПГ" in dep:
                vdpg_keys.update(dep["ВДПГ"].keys())
    return sorted(list(info_keys)), sorted(list(vdpg_keys))


def calculate_group_stats(data, cat_list, event_limit, info_keys):
    """Считает статистику только для ДОСТУПНЫХ депутатов группы"""
    p1_fact, p1_plan = 0, 0
    p2_fact, p2_plan = 0, 0

    available_count = 0
    for cat_name in cat_list:
        for dep in data.get(cat_name, []):
            if not dep.get("is_available"): continue  # Полностью исключаем из расчета

            available_count += 1
            # Индикатор 1 (Посты)
            posts = dep.get("Посты по информационным ударам", {})
            for k in info_keys:
                p1_plan += 1
                if posts.get(k, {}).get("score"): p1_fact += 1

            # Индикатор 2 (Мероприятия)
            p2_plan += event_limit
            ev_key = next((k for k in dep.keys() if "Мероприятия" in k), None)
            if ev_key:
                events = dep[ev_key]
                dep_fact = 0
                # Считаем обязательные
                for ek in ["1", "2", "3", "4 в рег. парламенте"][:event_limit]:
                    if events.get(ek, {}).get("score"): dep_fact += 1
                # Считаем опциональные
                for opt in events.get("опционально", []):
                    if opt.get("score"): dep_fact += 1

                # Прибавляем к группе, но не более лимита на депутата
                p2_fact += dep_fact

    pct1 = (p1_fact / p1_plan * 100) if p1_plan > 0 else 0
    pct2 = (p2_fact / p2_plan * 100) if p2_plan > 0 else 0
    return pct1, pct2


def calculate_metrics(data, info_keys, vdpg_keys):
    # Группа ЗС
    zs_p1, zs_p2 = calculate_group_stats(data, ["Депутаты Законодательных собраний регионов"], 4,
                                         info_keys)
    # Группа Остальные
    oth_p1, oth_p2 = calculate_group_stats(data, [
        "Депутаты административных центров регионов",
        "Депутаты муниципальных образований"
    ], 2, info_keys)

    # Среднее арифметическое между двумя группами (как вы и просили)
    ind1 = (zs_p1 + oth_p1) / 2
    ind2 = (zs_p2 + oth_p2) / 2

    total_k4 = (ind1 + ind2) / 2
    score_k4 = 0
    if total_k4 >= 75:
        score_k4 = 4
    elif total_k4 >= 50:
        score_k4 = 3
    elif total_k4 >= 25:
        score_k4 = 2
    elif total_k4 >= 10:
        score_k4 = 1

    # ВДПГ (К7) - Считается по всем доступным депутатам
    v_results = []
    for k in vdpg_keys:
        v_fact, v_plan = 0, 0
        all_cats = ["Депутаты Законодательных собраний регионов",
                    "Депутаты административных центров регионов",
                    "Депутаты муниципальных образований"]
        for cat in all_cats:
            for dep in data.get(cat, []):
                if not dep.get("is_available"): continue
                v_plan += 1
                if dep.get("ВДПГ", {}).get(k, {}).get("score"): v_fact += 1
        v_results.append((v_fact / v_plan * 100) if v_plan > 0 else 0)

    avg_v = sum(v_results) / len(v_results) if v_results else 0
    score_k7 = 2 if avg_v >= 75 else (1 if avg_v >= 50 else 0)

    return ind1, ind2, total_k4, score_k4, v_results, score_k7


def print_deputy_level(ws, data, row, level, level_color, info_keys, vdpg_keys):
    col_info_start = 6
    col_event_start = col_info_start + len(info_keys)
    col_vdpg_start = col_event_start + 15
    last_col = col_vdpg_start + len(vdpg_keys) - 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=level)
    c.fill = level_color
    style_cell(c, size=24, bold=True)
    ws.row_dimensions[row].height = 0.61 * INCH_TO_CHAR_H
    row += 1

    headers = ["", "ФИО", "Населенный пункт", "Контактный номер...", "Причины невзаимодействия..."]
    for i, h in enumerate(headers, 1):
        style_cell(ws.cell(row=row, column=i, value=h), size=18, bold=True)

    is_zs = "Законодательных" in level
    for i in range(15):
        col = col_event_start + i
        txt, f = "(опционально)", BEIGE_FILL
        if i < 4:
            f = ORANGE_FILL
            if is_zs:
                txt = str(i + 1) if i < 3 else "4\nв рег. парламенте"
            else:
                txt = str(i + 1) if i < 2 else "(опционально)"
                if i >= 2: f = BEIGE_FILL
        c = ws.cell(row=row, column=col, value=txt)
        c.fill = f
        style_cell(c, size=16, bold=True)

    ws.row_dimensions[row].height = 1.80 * INCH_TO_CHAR_H
    row += 1

    for i, dep in enumerate(data.get(level, []), 1):
        ws.cell(row=row, column=1, value=i)
        c_fio = ws.cell(row=row, column=2, value=dep.get('fio'))
        c_fio.fill = GREEN_FILL if dep.get('is_available') else RED_FILL
        style_cell(c_fio, size=18, horiz='left')

        if not dep.get('is_available'):
            row += 1;
            continue

        for col_idx, k in enumerate(['settlement', 'contact', 'reason'], 3):
            style_cell(ws.cell(row=row, column=col_idx, value=dep.get(k)), size=18, horiz='left')

        # Инфоудары
        posts = dep.get("Посты по информационным ударам", {})
        for idx, k in enumerate(info_keys):
            p = posts.get(k, {})
            c = ws.cell(row=row, column=col_info_start + idx, value=p.get('represent'))
            c.fill = GREEN_FILL if p.get('score') else RED_FILL
            style_cell(c, size=14, horiz='left', vert='top')

        # Мероприятия
        ev_key = next((k for k in dep.keys() if "Мероприятия" in k), None)
        events = dep.get(ev_key, {})
        limit = 4 if is_zs else 2
        for idx, ek in enumerate(["1", "2", "3", "4 в рег. парламенте"]):
            e = events.get(ek, {})
            if idx == limit:
                break
            c = ws.cell(row=row, column=col_event_start + idx, value=e.get('represent'))
            # if idx < limit:
            c.fill = GREEN_FILL if e.get('score') else RED_FILL
            # elif c.value and c.value != "Отсутствует ссылка":
            #     c.fill = GREEN_FILL if e.get('score') else RED_FILL
            style_cell(c, size=14, horiz='left', vert='top')

        opts = events.get("опционально", [])
        for idx in range(15 - limit):
            col = col_event_start + limit + idx
            if idx < len(opts):
                o = opts[idx]
                if o.get('represent') != "Отсутствует ссылка":
                    c = ws.cell(row=row, column=col, value=o.get('represent'))
                    c.fill = GREEN_FILL if o.get('score') else RED_FILL
                    style_cell(c, size=14, horiz='left', vert='top')

        # ВДПГ
        vdpg_d = dep.get("ВДПГ", {})
        for idx, k in enumerate(vdpg_keys):
            v = vdpg_d.get(k, {})
            c = ws.cell(row=row, column=col_vdpg_start + idx, value=v.get('represent'))
            c.fill = GREEN_FILL if v.get('score') else RED_FILL
            style_cell(c, size=14, horiz='left', vert='top')

        ws.row_dimensions[row].height = 1.35 * INCH_TO_CHAR_H
        row += 1
    return row


def generate_excel_bytes(data, info_titles=None, vdpg_titles=None):
    info_keys, vdpg_keys = get_dynamic_schema(data)
    num_info = len(info_keys)
    num_vdpg = len(vdpg_keys)
    ind1, ind2, total_k4, score_k4, v_res, score_k7 = calculate_metrics(data, info_keys, vdpg_keys)

    col_info_start, col_event_start = 6, 6 + num_info
    col_vdpg_start = col_event_start + 15
    total_cols = col_vdpg_start + num_vdpg - 1

    wb = Workbook()
    ws = wb.active

    # Ширина столбцов
    ws.column_dimensions['A'].width = 0.82 * INCH_TO_CHAR
    ws.column_dimensions['B'].width = 5.73 * INCH_TO_CHAR
    ws.column_dimensions['C'].width = 2.61 * INCH_TO_CHAR
    ws.column_dimensions['D'].width = 2.61 * INCH_TO_CHAR
    ws.column_dimensions['E'].width = 5.59 * INCH_TO_CHAR
    for i in range(col_info_start, col_event_start): ws.column_dimensions[
        get_column_letter(i)].width = 6.52 * INCH_TO_CHAR
    for i in range(col_event_start, col_vdpg_start): ws.column_dimensions[
        get_column_letter(i)].width = 2.43 * INCH_TO_CHAR
    for i in range(col_vdpg_start, total_cols + 1): ws.column_dimensions[
        get_column_letter(i)].width = 7.98 * INCH_TO_CHAR

    # Шапка
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c1 = ws.cell(1, 1, data.get('region'));
    c1.fill = BLUE_FILL
    style_cell(c1, size=28, bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c2 = ws.cell(2, 1, "Отчетный период: 21 февраля 2026 г. - 20 марта 2026 г.");
    c2.fill = BLUE_FILL
    style_cell(c2, size=26, bold=True)

    # Блок статистики
    stats = [
        ("% выполнения плана по Индикатору №1 «Посты по информационным ударам» Показателя №4",
         f"{ind1:.2f}%"),
        (
        "% выполнения плана по Индикатору №2 «Мероприятия по взаимодействию с избирателями...» Показателя №4",
        f"{ind2:.2f}%"),
        ("% выполнения Показателя №4", f"{total_k4:.2f}%"),
        ("Итоговый балл по Показателю №4", str(score_k4))
    ]
    for idx, pct in enumerate(v_res):
        stats.append((f"% выполнения приема {idx + 1}", f"{pct:.2f}%"))
    stats.append(("Итоговый балл по Показателю №7", str(score_k7)))

    for i, (lab, val) in enumerate(stats, 3):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
        cl, cv = ws.cell(i, 1, lab), ws.cell(i, 3, val)
        cl.fill = cv.fill = GREY_FILL
        style_cell(cl, size=18, bold=True)
        style_cell(cv, size=12, bold=True)
        ws.row_dimensions[i].height = 0.90 * INCH_TO_CHAR_H

    # Заголовки К4 и К7
    ws.merge_cells(start_row=3, start_column=col_info_start, end_row=5,
                   end_column=col_vdpg_start - 1)
    c4 = ws.cell(3, col_info_start,
                 "Показатель К4 «Работа депутатов с избирателями, отраслевыми сообществами»")
    c4.fill = PatternFill(start_color="adb9ca", fill_type="solid");
    style_cell(c4, size=28, bold=True)

    ws.merge_cells(start_row=3, start_column=col_vdpg_start, end_row=5, end_column=total_cols)
    c7 = ws.cell(3, col_vdpg_start, "Показатель К7\n«Всероссийский прием граждан (ВПГ)»")
    c7.fill = LIGHT_PURPLE_FILL;
    style_cell(c7, size=28, bold=True)

    # Групповые подзаголовки
    ws.merge_cells(start_row=6, start_column=col_info_start, end_row=6,
                   end_column=col_event_start - 1)
    cp = ws.cell(6, col_info_start, "Посты по информационным ударам")
    cp.fill = GREENY_FILL;
    style_cell(cp, size=36, bold=True)

    ws.merge_cells(start_row=6, start_column=col_event_start, end_row=9,
                   end_column=col_vdpg_start - 1)
    ce = ws.cell(6, col_event_start, "Мероприятия по взаимодействию с избирателями...")
    ce.fill = ORANGE_FILL;
    style_cell(ce, size=28, bold=True)

    for i in range(num_vdpg):
        col = col_vdpg_start + i
        ws.merge_cells(start_row=6, start_column=col, end_row=9, end_column=col)
        v_title = vdpg_titles[i] if vdpg_titles and i < len(vdpg_titles) else f"Прием {i + 1}"
        c = ws.cell(6, col, v_title);
        c.fill = LIGHT_PURPLE_FILL;
        style_cell(c, size=36, bold=True)

    for i in range(num_info):
        col = col_info_start + i
        ws.merge_cells(start_row=7, start_column=col, end_row=9, end_column=col)
        i_title = info_titles[i] if info_titles and i < len(
            info_titles) else f"Тема: {info_keys[i]}"
        c = ws.cell(7, col, i_title);
        c.fill = GREENY_FILL;
        style_cell(c, size=22, bold=True)

    # Отрисовка данных по уровням
    curr = 10
    levels = [("Депутаты Законодательных собраний регионов", YELLOW_FILL),
              ("Депутаты административных центров регионов", GREENY_FILL),
              ("Депутаты муниципальных образований", LIGHT_PURPLE_FILL)]
    for name, fill in levels: curr = print_deputy_level(ws, data, curr, name, fill, info_keys,
                                                        vdpg_keys)

    # Границы для всей таблицы
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in ws.iter_rows(max_row=curr - 1, max_col=total_cols):
        for cell in r: cell.border = border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)  # Возвращаем указатель в начало буфера

    return buffer

